#!/usr/bin/env python3
"""Build districts.geojson + blocks.geojson (with crop join) for deoria-bioregion."""
import json, re, unicodedata, difflib
import openpyxl
from shapely.geometry import shape, mapping
from shapely.ops import unary_union

SCRATCH = "/private/tmp/claude-501/-Users-mithunsheshagiri-work-lokaApps/eee9e18b-80b2-46a4-8f7b-e95a736de1f0/scratchpad"
OUT = "/Users/mithunsheshagiri/work/lokaApps/atlas/datasets/deoria-bioregion"
XLSX = "/Users/mithunsheshagiri/Downloads/Deoria_AttributeMapping.xlsx"

DISTRICTS = ["Deoria", "Kushi Nagar", "Gorakhpur"]           # bharatlas dtname
DISTRICT_LABEL = {"Kushi Nagar": "Kushinagar"}               # display name
CROPS = ["Sugarcane", "Banana", "Mustard", "Turmeric", "Rice", "Wheat", "Vegetable Pea", "Maize"]

def norm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z]", "", s.lower())

# block-name aliases: normalized xlsx/narrative name -> normalized bharatlas name
ALIAS = {
    "bishunpura": "vishunpura", "dudahi": "dudhahi", "kasia": "kasaya",
    "desahideoria": "desaideoria", "parthardeva": "pathardewa", "tarkulwa": "tarkalua",
    "kaptanganj": "kaptainganj", "nebuanaurangiya": "nebuanaurangia",
    "nebuanaurangia": "nebuanaurangia", "tamkuhi": "tamkuhiraj", "tamkuhiraj": "tamkuhiraj",
    "sewrahi": "seorahi", "kathkuiyanpadrauna": "padrauna", "kauriram": "kauriram",
    "desaideoria": "desaideoria",
}
def canon(name):
    n = norm(name)
    return ALIAS.get(n, n)

# ---------------------------------------------------------------- load xlsx
wb = openpyxl.load_workbook(XLSX, data_only=True)

# crop presence matrix
ws = wb["Crop details by block"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
crop_cols = {}
for i, h in enumerate(header):
    if h and str(h).strip() in CROPS:
        crop_cols[str(h).strip()] = i
matrix = {}   # canon(block) -> set(crops)
cur = None
for r in rows[1:]:
    if r[0]:
        cur = r[0]
    b = r[1]
    if not b:
        continue
    grown = {c for c, idx in crop_cols.items() if r[idx] is True}
    matrix.setdefault(canon(b), set()).update(grown)   # merge duplicate block rows

# narrative: context[(district,crop)] = details ; predominant blocks
wsn = wb["Block wise crop details"]
CROP_NARR = {"Rice (Paddy)": ["Rice"], "Rice & Wheat": ["Rice", "Wheat"]}
context = {}      # canon(block) -> list of {title, body, predominant}
crop_context_by_district = {}  # (districtLabel, crop) -> body
predominant = {}  # canon(block) -> set(crops)
for r in list(wsn.iter_rows(values_only=True))[1:]:
    if not r or not r[0]:
        continue
    dist, crop, blocks, details = r[0], r[1], r[2], r[3]
    crops = CROP_NARR.get(str(crop).strip(), [str(crop).strip()])
    for c in crops:
        crop_context_by_district[(str(dist).strip(), c)] = (str(details).strip() if details else "")
        for b in re.split(r"[,\n]", str(blocks or "")):
            b = re.sub(r"\(.*?\)", "", b).strip()
            if not b:
                continue
            predominant.setdefault(canon(b), set()).add(c)

# district-level general info
wsg = wb["General information"]
g = {}
for r in wsg.iter_rows(values_only=True):
    if r and r[0]:
        g[str(r[0]).strip()] = [c for c in r[1:]]
# columns order = Deoria, Khushinagar, Gorakhpur
GEN_ORDER = ["Deoria", "Kushinagar", "Gorakhpur"]
def gen_val(key, dlabel):
    idx = GEN_ORDER.index(dlabel)
    row = g.get(key, [])
    return str(row[idx]).strip() if idx < len(row) and row[idx] else ""

# ---------------------------------------------------------------- districts
src = json.load(open(f"{SCRATCH}/bl_districts.geojson"))
dfeats = []
for f in src["features"]:
    p = f["properties"]
    if "UTTAR" in str(p.get("stname", "")).upper() and p.get("dtname") in DISTRICTS:
        dt = p["dtname"]; label = DISTRICT_LABEL.get(dt, dt)
        geom = shape(f["geometry"]).buffer(0).simplify(0.0007, preserve_topology=True)
        dfeats.append({"type": "Feature",
            "properties": {
                "name": label, "kind": "district", "dist_lgd": p.get("dist_lgd"),
                "crops": gen_val("Crops", label),
                "soil": gen_val("Soil ", label) or gen_val("Soil", label),
                "water": gen_val("Water resources (Major rivers/ lakes)", label),
                "odop": gen_val("ODOP", label),
                "valuechain": gen_val("Value chain infra (local mandis, sugar mills, rice mills, flour processing units.. etc)", label),
            },
            "geometry": mapping(geom)})
json.dump({"type": "FeatureCollection", "features": dfeats},
          open(f"{OUT}/districts.geojson", "w"))
print(f"districts.geojson: {len(dfeats)} features")

# ---------------------------------------------------------------- blocks
src = json.load(open(f"{SCRATCH}/bl_blocks.geojson"))
bfeats = []
unmatched = []
all_canon = set(matrix) | set(predominant)
bl_keys = []
for f in src["features"]:
    p = f["properties"]
    if "UTTAR" not in str(p.get("state", "")).upper() or p.get("district") not in DISTRICTS:
        continue
    dt = p["district"]; label = DISTRICT_LABEL.get(dt, dt)
    raw = str(p.get("block_name", "")).title()
    key = canon(raw)
    bl_keys.append(key)
    grown = sorted(matrix.get(key, set()), key=CROPS.index)
    predom = predominant.get(key, set())
    notes = []
    for c in grown:
        body = crop_context_by_district.get((label, c), "") or crop_context_by_district.get((dt, c), "")
        notes.append({"title": c + (" — predominant" if c in predom else ""),
                      "body": body, "predominant": c in predom})
    geom = shape(f["geometry"]).buffer(0).simplify(0.0004, preserve_topology=True)
    bfeats.append({"type": "Feature",
        "properties": {
            "name": raw, "district": label, "kind": "block",
            "block_lgd": p.get("block_lgd"),
            "crops": grown, "predominantFor": sorted(predom, key=lambda x: CROPS.index(x) if x in CROPS else 99),
            "cropNotes": notes,
            "ncrops": len(grown),
        },
        "geometry": mapping(geom)})
json.dump({"type": "FeatureCollection", "features": bfeats},
          open(f"{OUT}/blocks.geojson", "w"))
print(f"blocks.geojson: {len(bfeats)} features")

# validation: which matrix/predominant keys never matched a polygon
bl_set = set(bl_keys)
miss_matrix = [k for k in matrix if k not in bl_set]
miss_predom = [k for k in predominant if k not in bl_set]
print("matrix keys not matched to a polygon:", miss_matrix)
print("predominant keys not matched:", miss_predom)
# per-crop block counts (sanity)
from collections import Counter
cnt = Counter()
for f in bfeats:
    for c in f["properties"]["crops"]:
        cnt[c]+=1
print("crop -> #blocks:", dict(cnt))
# spot checks
for chk in ["Baitalpur", "Belghat", "Deoria Sadar", "Padrauna", "Pipraich"]:
    for f in bfeats:
        if f["properties"]["name"].lower()==chk.lower():
            print(f"  {chk} ({f['properties']['district']}): {f['properties']['crops']}")
